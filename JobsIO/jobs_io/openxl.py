'''
Handles setting cells, saving workbook.
'''
import openpyxl

output_name = "job_listing.xlsx"


def add_to_workbook(paired_list: list[tuple]):
    wb = openpyxl.Workbook()
    
    sheet = wb.active

    # Write headers
    sheet['A1'] = 'Job ID'
    sheet['B1'] = 'Salary Range'
    sheet['C1'] = 'Link'

    for row_num, (job_id, salary_range, job_link) in enumerate(paired_list, start=2):
        sheet.cell(row=row_num, column=1, value=job_id)
        sheet.cell(row=row_num, column=2, value=salary_range)
        sheet.cell(row=row_num, column=3, value=job_link)

    wb.save(output_name)